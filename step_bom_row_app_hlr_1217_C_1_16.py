# step_bom_inline_edit_hlr_toolinglist_v13_fix_height_final.py

import tempfile
from io import BytesIO
import base64
import json
from datetime import datetime
from uuid import uuid4
import html as pyhtml
import re
import math

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import matplotlib.pyplot as plt
from matplotlib.collections import LineCollection

from OCP.STEPControl import STEPControl_Reader
# ---- BRepGProp (Volume) ----
def _volume_properties(shape, props):
    """Call OpenCascade volume properties in a way that works for both
    pythonocc-core and cadquery-ocp (OCP).

    Different wheels expose different symbols (e.g. brepgprop_VolumeProperties,
    VolumeProperties, VolumeProperties_1 ...). We discover and call what exists.
    """
    # 1) pythonocc-core style wrapper
    try:
        from OCP.BRepGProp import brepgprop_VolumeProperties as _fn
        _fn(shape, props)
        return
    except Exception:
        pass

    # 2) cadquery-ocp / OCP style: discover any callable containing 'VolumeProperties'
    import OCP.BRepGProp as _bg

    for name, obj in vars(_bg).items():
        if "VolumeProperties" in name and callable(obj):
            try:
                obj(shape, props)
                return
            except Exception:
                continue

    cls = getattr(_bg, "BRepGProp", None)
    if cls is not None:
        for name, obj in vars(cls).items():
            if "VolumeProperties" in name and callable(obj):
                try:
                    obj(shape, props)
                    return
                except Exception:
                    continue

    raise ImportError("Cannot find a usable VolumeProperties function in OCP.BRepGProp")

from OCP.GProp import GProp_GProps
from OCP.gp import gp_Dir, gp_Pnt, gp_Ax2
from OCP.HLRAlgo import HLRAlgo_Projector
from OCP.HLRBRep import HLRBRep_Algo, HLRBRep_HLRToShape
from OCP.TopExp import TopExp_Explorer
from OCP.TopAbs import TopAbs_EDGE
from OCP.BRepAdaptor import BRepAdaptor_Curve
from OCP.Bnd import Bnd_Box
try:
    from OCP.Bnd import Bnd_OBB
except Exception:
    Bnd_OBB = None
# ---- BRepBndLib (Bounding) ----
# cadquery-ocp / pythonocc-core have slightly different symbol exports.
# We provide robust wrappers that try multiple candidate callables at runtime.

def _pick_callables(module_or_cls, name_hints):
    out = []
    for n in name_hints:
        obj = getattr(module_or_cls, n, None)
        if callable(obj):
            out.append(obj)
    # also scan for anything containing 'Add' / 'OBB'
    try:
        for n in dir(module_or_cls):
            if any(h in n for h in ("Add", "OBB")):
                obj = getattr(module_or_cls, n, None)
                if callable(obj):
                    out.append(obj)
    except Exception:
        pass
    # de-dup while keeping order
    seen = set()
    uniq = []
    for f in out:
        fid = id(f)
        if fid not in seen:
            uniq.append(f)
            seen.add(fid)
    return uniq


def brepbndlib_Add(shape, box):
    """Add shape to Bnd_Box (AABB)."""
    last_err = None
    try:
        # pythonocc-core style
        from OCP.BRepBndLib import brepbndlib_Add as _f
        return _f(shape, box)
    except Exception as e:
        last_err = e

    try:
        import OCP.BRepBndLib as _m
    except Exception as e:
        raise ImportError("Cannot import OCP.BRepBndLib") from e

    cands = []
    cands += _pick_callables(_m, ("brepbndlib_Add", "Add", "Add_1", "Add_2", "add"))
    cls = getattr(_m, "BRepBndLib", None)
    if cls is not None:
        cands += _pick_callables(cls, ("Add", "Add_1", "Add_2", "add"))

    for f in cands:
        try:
            return f(shape, box)
        except Exception as e:
            last_err = e
            continue

    raise AttributeError("No compatible BRepBndLib Add(...) found") from last_err


def brepbndlib_AddOBB(shape, obb, *args):
    """Add shape to Bnd_OBB when available."""
    last_err = None
    try:
        from OCP.BRepBndLib import brepbndlib_AddOBB as _f
        return _f(shape, obb, *args)
    except Exception as e:
        last_err = e

    try:
        import OCP.BRepBndLib as _m
    except Exception:
        return None

    cands = []
    cands += _pick_callables(_m, ("brepbndlib_AddOBB", "AddOBB", "AddOBB_1", "AddOBB_2"))
    cls = getattr(_m, "BRepBndLib", None)
    if cls is not None:
        cands += _pick_callables(cls, ("AddOBB", "AddOBB_1", "AddOBB_2"))

    for f in cands:
        try:
            return f(shape, obb, *args)
        except TypeError:
            # try without extra args
            try:
                return f(shape, obb)
            except Exception as e2:
                last_err = e2
                continue
        except Exception as e:
            last_err = e
            continue

    return None


# ---------------- UI knobs ----------------
IMAGE_WIDTH_PX = 85
HEADER_HEIGHT_PX = 44

# ---------------- Material density (g/cm^3) ----------------
MATERIAL_DENSITY = {
    "PC": 1.20,
    "ABS": 1.05,
    "PC/ABS": 1.15,
    "PC+GF10%": 1.27,
    "PC+GF20%": 1.34,
    "PC+GF30%": 1.42,
    "ADC 12": 2.70,
    "PP": 0.90,
    "PA": 1.14,
    "POM": 1.41,
    "Rubber": 1.10,
    "Silicon": 1.12,  # keep Silicon
}
CATEGORY_OPTIONS = [
    "Sub-Assy",
    "Double Shot Assy",
    "plastic molding (single shot)",
    "plastic molding (double shot, 1st shot)",
    "plastic molding (double shot, 2nd shot)",
    "plastic molding (insert molding)",
    "diecasting",
    "sheet metal",
    "Al extrusion",
    "MIM",
    "heat pressing (rubber, silicon)",
    "die cut",
    "P+R (P)",
    "P+R (R)",
]

TOOLING_MATERIAL_OPTIONS = ["", "S136", "NAK80", "P20", "Other"]
RUNNER_OPTIONS = ["hot runner", "cold runner", "N/A"]

ASSEMBLY_OPTIONS = ["螺絲鎖附", "熱熔", "點膠", "雙面膠黏貼", "超音波", "緊配組裝", "其他"]


# ---------------- state ----------------
def ensure_state():
    if "bom" not in st.session_state:
        st.session_state["bom"] = []
    if "project" not in st.session_state:
        st.session_state["project"] = ""
    if "engineer" not in st.session_state:
        st.session_state["engineer"] = ""
    if "uploader_key" not in st.session_state:
        st.session_state["uploader_key"] = str(uuid4())
    if "new_bom_used" not in st.session_state:
        st.session_state["new_bom_used"] = False
    if "bom_context_ready" not in st.session_state:
        st.session_state["bom_context_ready"] = False
    if "bom_version" not in st.session_state:
        st.session_state["bom_version"] = 0
    if "tooling_version" not in st.session_state:
        st.session_state["tooling_version"] = 0
    if "add_mode" not in st.session_state:
        st.session_state["add_mode"] = ""  # "", "part", "subassy"
    if "subassy_parent_id" not in st.session_state:
        st.session_state["subassy_parent_id"] = ""
    if "subassy_parent_seq" not in st.session_state:
        st.session_state["subassy_parent_seq"] = ""
    if "subassy_child_n" not in st.session_state:
        st.session_state["subassy_child_n"] = 0
    if "subassy_child_meta" not in st.session_state:
        st.session_state["subassy_child_meta"] = {}  # child_seq -> {method, qty, other}
    if "subassy_child_uploaded" not in st.session_state:
        st.session_state["subassy_child_uploaded"] = {}  # child_seq -> bool
    if "subassy_child_files" not in st.session_state:
        st.session_state["subassy_child_files"] = {}  # child_seq -> filename
    if "subassy_uploader_key" not in st.session_state:
        st.session_state["subassy_uploader_key"] = str(uuid4())

    if "part_pending_files" not in st.session_state:
        st.session_state["part_pending_files"] = []  # list of {id,name,size,bytes}


    if "texture_edit_id" not in st.session_state:
        st.session_state["texture_edit_id"] = ""
    if "texture_text" not in st.session_state:
        st.session_state["texture_text"] = ""
    if "process_edit_id" not in st.session_state:
        st.session_state["process_edit_id"] = ""
    if "process_text" not in st.session_state:
        st.session_state["process_text"] = ""


def reset_bom():
    st.session_state["bom"] = []


def esc(s: str) -> str:
    return pyhtml.escape(s or "", quote=True)


def esc_multiline(s: str) -> str:
    return esc(s).replace("\n", "<br>")


def png_bytes_to_b64(png_bytes: bytes) -> str:
    return base64.b64encode(png_bytes).decode("utf-8")



# ---------------- Excel export ----------------
def _safe_filename(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return "Project"
    return re.sub(r'[\\/:*?"<>|]+', "_", s)


def build_bom_export_df(items: list[dict]) -> pd.DataFrame:
    rows = []
    for i, it in enumerate(items, start=1):
        rows.append({
            "#": it.get("seq", str(i)),
            "Part No.": str(it.get("part_no", "TBD")).upper(),
            "Part Name": it.get("part_name", ""),
            "Preview": "",
            "QTY": int(it.get("qty", 1)),
            "Category": it.get("category", ""),
            "Material": it.get("material", ""),
            "Resin Code": it.get("material_spec", ""),
            "Texture": it.get("texture", ""),
            "2nd Process / Assembly": it.get("second_process", ""),
            "Density": float(it.get("density", 0.0)),
            "Vol (cm^3)": float(it.get("volume_cm3", 0.0)),
            "Weight (g)": float(it.get("mass_g", 0.0)),
            "_preview_b64": it.get("preview_png_b64", ""),
        })
    return pd.DataFrame(rows)


def build_tooling_export_df(items: list[dict]) -> pd.DataFrame:
    rows = []
    for i, it in enumerate(items, start=1):
        if str(it.get("category","")).lower() == "sub-assy":
            continue
        rows.append({
            "#": it.get("seq", str(i)),
            "Part No.": str(it.get("part_no", "TBD")).upper(),
            "Tooling No.": it.get("tooling_no", "TBD") or "TBD",
            "Part Name": it.get("part_name", ""),
            "Preview": "",
            "Category": it.get("category", ""),
            "Runner": it.get("runner", "N/A"),
            "Tooling Life": it.get("tooling_life_k", ""),
            "Cavities": it.get("cavities", "1"),
            "Texture": it.get("texture", ""),
            "Cavity Mat.": it.get("cavity_material", ""),
            "Core Mat.": it.get("core_material", ""),
            "Lifter": int(it.get("lifter_qty", 0)),
            "Slider": int(it.get("slider_qty", 0)),
            "_preview_b64": it.get("preview_png_b64", ""),
        })
    return pd.DataFrame(rows)


def _has_openpyxl() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except Exception:
        return False


def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    from io import BytesIO
    import base64

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils.units import pixels_to_EMU
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
    except Exception:
        raise ModuleNotFoundError(
            "openpyxl is required to export .xlsx. Install it with: pip install openpyxl"
        )

    out = BytesIO()

    visible_cols = [c for c in df.columns if c != "_preview_b64"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True)

    for j, col in enumerate(visible_cols, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    data_df = df[visible_cols].copy()
    for i, row in enumerate(data_df.itertuples(index=False, name=None), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    ws.freeze_panes = "A2"

    def _col_width_to_pixels(width: float) -> int:
        return int(width * 7 + 5)

    fixed_width = {
        "Preview": 14,
    }
    max_cap_default = 36
    max_cap_special = {
        "Part Name": 42,
        "2nd Process / Assembly": 32,
        "Remark": 42,
    }
    for j, col in enumerate(visible_cols, start=1):
        letter = get_column_letter(j)
        if col in fixed_width:
            ws.column_dimensions[letter].width = fixed_width[col]
            continue

        max_len = len(str(col)) if col is not None else 8
        for v in data_df[col].tolist() if col in data_df.columns else []:
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)

        cap = max_cap_special.get(col, max_cap_default)
        width = min(max_len + 2, cap)
        width = max(width, 10)
        ws.column_dimensions[letter].width = width

    preview_col_idx = None
    if "Preview" in visible_cols and "_preview_b64" in df.columns:
        preview_col_idx = visible_cols.index("Preview") + 1
        for i in range(2, len(data_df) + 2):
            ws.row_dimensions[i].height = 60

    left_cols = {"Part Name", "2nd Process / Assembly"}
    for j, col in enumerate(visible_cols, start=1):
        align = left_align if col in left_cols else center_align
        for i in range(2, len(data_df) + 2):
            ws.cell(row=i, column=j).alignment = align

    numeric_int_cols = {"Cavities", "QTY", "#", "Lifter", "Slider", "Tooling Life"}
    for j, col in enumerate(visible_cols, start=1):
        if col in numeric_int_cols:
            for i in range(2, len(data_df) + 2):
                c = ws.cell(row=i, column=j)
                v = c.value
                if v is None:
                    continue
                try:
                    if isinstance(v, str):
                        vv = v.strip()
                        if vv == "":
                            continue
                        if vv.isdigit():
                            c.value = int(vv)
                        else:
                            fv = float(vv)
                            c.value = int(fv) if fv.is_integer() else fv
                    elif isinstance(v, float) and float(v).is_integer():
                        c.value = int(v)
                    if isinstance(c.value, int):
                        c.number_format = "0"
                except Exception:
                    pass

    num3_cols = {"Density", "Vol (cm^3)"}
    num2_cols = {"Unit Weight (g)", "Total Weight (g)", "Total Weight (kg)", "Unit Price (USD)", "Tooling Price (USD)"}
    for j, col in enumerate(visible_cols, start=1):
        if col in num3_cols:
            for i in range(2, len(data_df) + 2):
                ws.cell(row=i, column=j).number_format = "0.000"
        if col in num2_cols:
            for i in range(2, len(data_df) + 2):
                ws.cell(row=i, column=j).number_format = "0.00"

    if preview_col_idx is not None:
        col_letter = get_column_letter(preview_col_idx)
        col_px = _col_width_to_pixels(ws.column_dimensions[col_letter].width or 14)

        for r in range(len(df)):
            b64 = df.iloc[r].get("_preview_b64", None)
            if not b64:
                continue
            try:
                img_bytes = base64.b64decode(b64)
                img_io = BytesIO(img_bytes)
                img = XLImage(img_io)
                img.height = 50
                img.width = 50

                row_idx = r + 2
                row_px = int((ws.row_dimensions[row_idx].height or 60) * 96 / 72)

                off_x = max(int((col_px - img.width) / 2), 0)
                off_y = max(int((row_px - img.height) / 2), 0)

                marker = AnchorMarker(
                    col=preview_col_idx - 1, colOff=pixels_to_EMU(off_x),
                    row=row_idx - 1, rowOff=pixels_to_EMU(off_y),
                )
                size = XDRPositiveSize2D(pixels_to_EMU(img.width), pixels_to_EMU(img.height))
                img.anchor = OneCellAnchor(_from=marker, ext=size)

                ws.add_image(img)
            except Exception:
                pass

    wb.save(out)
    return out.getvalue()


def recalc_item(item: dict) -> dict:
    """Recalculate weight from *editable* volume + density.

    - volume_cm3: keep the original / edited total volume (cm^3)
    - l_mm/w_mm/h_mm: only for BOM read-only display (L x W x H), DO NOT override volume_cm3
    """
    vol = float(item.get("volume_cm3", item.get("step_volume_cm3", 0.0)) or 0.0)
    item["volume_cm3"] = vol
    den = float(item.get("density", 0.0) or 0.0)
    item["mass_g"] = vol * den
    return item

# ---------------- Cavities suggestion ----------------
# Heuristic rule (no external DB): use bounding-box size (L/W/H) & category to suggest cavities.
# - Normal molding/diecast: show as 1xN (e.g., 1x1, 1x2, 1x4...)
# - Double-shot: show as N+N (e.g., 1+1, 2+2...)
# Users can still override in the editor.


# ---- Rubber / Silicon heat pressing cavities helper ----
# Default platen size (mm). Adjust if your press uses different platen dimensions.
RUBBER_PLATEN_X_MM = 400
RUBBER_PLATEN_Y_MM = 400

def can_make_1x8_rubber_press(L, W, H, platen_x, platen_y, pitch_margin=10, edge_margin=30):
    # geometry pitch
    px = L + pitch_margin
    py = W + pitch_margin

    # usable platen
    ux = platen_x - 2*edge_margin
    uy = platen_y - 2*edge_margin
    if ux <= 0 or uy <= 0:
        return False, 0

    # two orientations
    nx1 = math.floor(ux / px)
    ny1 = math.floor(uy / py)
    cap1 = nx1 * ny1

    nx2 = math.floor(ux / py)
    ny2 = math.floor(uy / px)
    cap2 = nx2 * ny2

    capacity = max(cap1, cap2)

    # thickness guard (conservative)
    if H >= 45:
        return False, capacity

    return capacity >= 8, capacity




def get_high_capacity_cavity(l_mm, w_mm, h_mm):
    """High-capacity cavities suggestion for heat pressing (rubber, silicon).

    Rule provided by user (based on projected area A=L*W and thickness H).
    Returns N as string.
    """
    try:
        dims = sorted([float(l_mm or 0.0), float(w_mm or 0.0), float(h_mm or 0.0)], reverse=True)
        L, W, H = dims[0], dims[1], dims[2]
        A = L * W  # projected area (mm^2)

        if H >= 100 or A >= 45000:      # 特大件
            n = 2
        elif H >= 90 or A >= 25000:     # 大件
            n = 4
        elif H >= 80 or A >= 12000:     # 中大件
            n = 8
        elif H >= 60 or A >= 3500:
            n = 16
        elif H >= 25 or A >= 1500:
            n = 32
        elif H >= 5 or A >= 500:        # 小件
            n = 64
        else:
            n = 100                     # 微型件
    except Exception:
        n = 0

    return str(n)

def effective_thickness_from_volume(l_mm: float, w_mm: float, h_mm: float, vol_cm3: float, k: float = 1.25):
    """Compute an 'effective thickness' for thin parts to avoid local features (boss/rib) dominating H.

    Strategy:
    - Treat the two largest bbox dims as L/W footprint, smallest as Hmin
    - Estimate thickness from volume / projected area: t_est = V / (L*W)
      where V in mm^3, L*W in mm^2  -> t_est in mm
    - Use H_eff = min(Hmin, t_est * k) as a conservative stable thickness
    Returns: (L, W, H_eff, Hmin, t_est)
    """
    try:
        dims = sorted([float(l_mm or 0.0), float(w_mm or 0.0), float(h_mm or 0.0)], reverse=True)
        L, W, Hmin = dims[0], dims[1], dims[2]
    except Exception:
        L = W = Hmin = 0.0

    A = max(L * W, 1e-6)
    V_mm3 = float(vol_cm3 or 0.0) * 1000.0
    t_est = V_mm3 / A if V_mm3 > 0 else 0.0
    H_eff = min(Hmin, t_est * float(k)) if (Hmin > 0 and t_est > 0) else Hmin
    return L, W, H_eff, Hmin, t_est


def suggest_cavities_str(category: str, l_mm: float, w_mm: float, h_mm: float, vol_cm3: float) -> str:
    """Suggest cavities string.

    Display rule (per your latest spec): always show N only (no '1xN', no 'N+N').

    Key improvement for thin solid parts:
    - For non-heat-press categories, use an 'effective thickness' derived from Volume / Footprint area to avoid
      local features (boss/rib) inflating the thickness decision.
      H_eff = min(Hmin, (V/(L*W)) * k), where k defaults to 1.25.

    Special rule:
    - If Category == 'heat pressing (rubber, silicon)', use the user-provided high-capacity rule (get_high_capacity_cavity),
      based on bbox L/W/H (no volume correction).
    """
    cat_raw = (category or "").strip()
    cat = cat_raw.lower()

    # Not applicable
    if cat in {"sub-assy", "sub assy", "subassy"}:
        return "-"

    # Only apply to processes that typically have tooling cavities; keep simple.
    if not (
        cat.startswith("plastic molding")
        or cat in {"diecasting", "mim"}
        or cat.startswith("p+r")
        or cat == "heat pressing (rubber, silicon)"
        or cat in {"sub-assy child", "sub assy child", "subassy child"}
    ):
        return "-"

    # footprint + thickness (thin-part robust)
    L, W, H_eff, Hmin, t_est = effective_thickness_from_volume(l_mm, w_mm, h_mm, vol_cm3, k=1.25)

    # ---- Special: heat pressing (rubber, silicon) ----
    if cat == "heat pressing (rubber, silicon)":
        # Use the bbox-based rule exactly as specified (no volume correction)
        return get_high_capacity_cavity(L, W, Hmin)

    # ---- Original heuristic for other categories (using H_eff) ----
    A = L * W  # projected area (mm^2)

    if H_eff >= 50 or A >= 6000:
        n = 1
    elif H_eff >= 25 or A >= 3000:
        n = 2
    elif H_eff >= 12 or A >= 1200:
        n = 4
    else:
        n = 4  # conservative default for very small parts

    return str(n)
def get_next_main_seq(bom_items: list[dict]) -> str:
    """Return next main sequence number (ignore child seq like 5-1, 5-2)."""
    max_main = 0
    for it in bom_items:
        seq = str(it.get("seq", "")).strip()
        if "-" in seq:
            continue
        try:
            n = int(seq)
            if n > max_main:
                max_main = n
        except ValueError:
            continue
    return str(max_main + 1)


# ---------------- STEP / geometry ----------------
def load_shape_from_step(path: str):
    reader = STEPControl_Reader()
    status = reader.ReadFile(path)
    if status != 1:
        raise RuntimeError(f"STEP read failed, status={status}")
    reader.TransferRoots()
    return reader.OneShape()


def compute_volume_mm3(shape) -> float:
    props = GProp_GProps()
    _volume_properties(shape, props)
    return float(props.Mass())


def compute_bbox_dims_mm(shape) -> tuple[float, float, float]:
    """Return bounding-box dimensions (dx, dy, dz) in mm.

    Prefer Oriented Bounding Box (OBB) when available to reduce over-estimation for rotated parts.
    Fallback to Axis-Aligned Bounding Box (AABB).
    """
    # 1) OBB
    try:
        if (Bnd_OBB is not None) and (brepbndlib_AddOBB is not None):
            obb = Bnd_OBB()
            # (use_triangulation, is_optimal, is_triangulation) flags vary by OCC build
            # We keep a permissive call signature via try.
            try:
                brepbndlib_AddOBB(shape, obb, True, True, True)
            except TypeError:
                brepbndlib_AddOBB(shape, obb)
            dx = float(obb.XHSize() * 2.0)
            dy = float(obb.YHSize() * 2.0)
            dz = float(obb.ZHSize() * 2.0)
            dx, dy, dz = max(0.0, dx), max(0.0, dy), max(0.0, dz)
            if dx > 0 and dy > 0 and dz > 0:
                return dx, dy, dz
    except Exception:
        pass

    # 2) AABB fallback
    box = Bnd_Box()
    brepbndlib_Add(shape, box)
    xmin, ymin, zmin, xmax, ymax, zmax = box.Get()
    dx = max(0.0, float(xmax - xmin))
    dy = max(0.0, float(ymax - ymin))
    dz = max(0.0, float(zmax - zmin))
    return dx, dy, dz



# ---------------- HLR render ----------------
def make_hlr_png_bytes(
    shape,
    dpi=200,
    fig_size=4,
    outline_width=1.8,
    edge_width=0.9,
    color="#2f2f2f",
    sample_n=60,
) -> bytes:
    view_dir = gp_Dir(1, -1, 1)
    projector = HLRAlgo_Projector(gp_Ax2(gp_Pnt(0, 0, 0), view_dir))

    algo = HLRBRep_Algo()
    algo.Add(shape)
    algo.Projector(projector)
    algo.Update()
    algo.Hide()

    hlr = HLRBRep_HLRToShape(algo)
    layers = [(hlr.OutLineVCompound(), outline_width), (hlr.VCompound(), edge_width)]

    fig, ax = plt.subplots(figsize=(fig_size, fig_size))
    any_line = False

    for comp, lw in layers:
        lines = []
        exp = TopExp_Explorer(comp, TopAbs_EDGE)
        while exp.More():
            edge = exp.Current()
            adaptor = BRepAdaptor_Curve(edge)
            first = float(adaptor.FirstParameter())
            last = float(adaptor.LastParameter())
            if not (last > first):
                exp.Next()
                continue

            pts = []
            for i in range(sample_n + 1):
                t = first + (last - first) * i / sample_n
                p = adaptor.Value(t)
                pts.append([p.X(), p.Y()])

            if len(pts) >= 2:
                lines.append(pts)
                any_line = True
            exp.Next()

        if lines:
            ax.add_collection(LineCollection(lines, colors=color, linewidths=lw))

    if not any_line:
        raise RuntimeError("No HLR edges generated")

    ax.autoscale()
    ax.set_aspect("equal")
    ax.axis("off")

    buf = BytesIO()
    plt.savefig(buf, dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


# ---------------- HTML tables ----------------
def _render_html_table(title: str, columns: list[tuple[str, str]], rows: list[list[str]], row_h_px: int, header_h_px: int):
    """Render a read-only HTML table with STEP preview images."""
    img_max_h = max(120, int(row_h_px * 0.78))

    ths = "".join([f'<th style="width:{w};">{esc(h)}</th>' for h, w in columns])

    trs = ""
    for r in rows:
        tds = ""
        for cell in r:
            tds += f"""
            <td style="
                border-bottom:1px solid #eee;
                padding:10px 10px;
                height:{row_h_px}px;
                text-align:center;
                vertical-align:middle;
                overflow:hidden;
                word-break:break-word;
            ">{cell}</td>
            """
        trs += f"<tr>{tds}</tr>"

    html = """
    <html>
    <head>
      <style>
        body {{ margin:0; padding:0; }}
        .wrap {{
          width:100%;
          overflow-x:auto;
          padding-bottom:6px;
          margin-bottom:0px;
        }}
        .wrap table {{
          width:100%;
          min-width:980px;
          border-collapse:collapse;
          table-layout:fixed;
          font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,"Noto Sans",Arial,"PingFang TC","Microsoft JhengHei",sans-serif;
        }}
        .wrap th {{
          padding:10px;
          height:{header_h_px}px;
          border-bottom:2px solid #ddd;
          text-align:center;
          font-weight:700;
          font-size:12px;
        }}
        .wrap td {{ font-size:12px; }}
        .imgbox {{
          height:{row_h_px}px;
          display:flex;
          align-items:center;
          justify-content:center;
        }}
        .wrap img {{
          width:{IMAGE_WIDTH_PX}px;
          max-height:{img_max_h}px;
          height:auto;
          display:block;
          object-fit:contain;
        }}
        .mono {{ font-variant-numeric: tabular-nums; }}
        .ml {{
          line-height:1.2;
          max-height:{row_h_px_minus}px;
          overflow:hidden;
        }}
      </style>

      <script>
        function _docHeight() {{
          const de = document.documentElement;
          const b = document.body;
          return Math.max(
            de ? de.scrollHeight : 0,
            b ? b.scrollHeight : 0,
            de ? de.offsetHeight : 0,
            b ? b.offsetHeight : 0
          );
        }}

        function _postSetHeight(h) {{
          try {{
            window.parent.postMessage({{ isStreamlitMessage: true, type: "streamlit:setFrameHeight", height: h }}, "*");
          }} catch (e) {{}}
        }}

        function _setHeight() {{
          const h = _docHeight() + 8;
          _postSetHeight(h);
        }}

        function _kick() {{
          _setHeight();
          let n = 0;
          const timer = setInterval(() => {{
            _setHeight();
            n++;
            if (n > 12) clearInterval(timer);
          }}, 200);

          try {{
            const ro = new ResizeObserver(() => _setHeight());
            ro.observe(document.body);
            if (document.documentElement) ro.observe(document.documentElement);
          }} catch (e) {{}}

          document.querySelectorAll("img").forEach(img => {{
            img.addEventListener("load", _setHeight);
            img.addEventListener("error", _setHeight);
          }});
        }}

        window.addEventListener("load", _kick);
        window.addEventListener("resize", () => {{
          _setHeight();
          setTimeout(_setHeight, 120);
        }});
      </script>
    </head>

    <body>
      <div class="wrap">
        <table>
          <thead><tr>{ths}</tr></thead>
          <tbody>{trs}</tbody>
        </table>
      </div>
    </body>
    </html>
    """.format(
        header_h_px=header_h_px,
        row_h_px=row_h_px,
        IMAGE_WIDTH_PX=IMAGE_WIDTH_PX,
        img_max_h=img_max_h,
        row_h_px_minus=max(30, row_h_px - 18),
        ths=ths,
        trs=trs,
    )

    st.subheader(title)

    # ✅ BUG FIX v13: 
    # 之前只加固定緩衝，但每一列有 padding:20px 造成累積誤差
    # 這裡用『padding + buffer』模型避免累積誤差（不再寫死 30px）
    # 確保筆數變多時，誤差不會累積導致切到底部

    # Height buffer tuning:


    # - ROW_PADDING_Y: matches td padding top+bottom (10px + 10px)


    # - ROW_EXTRA_BUF: extra safety for font/line-height/image decode variance


    ROW_PADDING_Y = 20


    ROW_EXTRA_BUF = 10


    total_height = header_h_px + len(rows) * (row_h_px + ROW_PADDING_Y + ROW_EXTRA_BUF) + 50
    components.html(html, height=total_height, scrolling=False)


def render_bom_readonly(items: list[dict]) -> None:
    row_h = max(150, int(IMAGE_WIDTH_PX * 1.7))

    columns = [
        ("#", "4%"),
        ("Part No.", "7%"),
        ("Part Name", "13%"),
        ("Preview", "12%"),
        ("QTY", "5%"),
        ("Category", "13%"),
        ("Material", "7%"),
        ("Resin Code", "11%"),
        ("Texture", "10%"),
        ("2nd Process / Assembly", "14%"),
        ("Vol (mm)", "6%"),
        ("Weight (g)", "5%"),
    ]

    rows = []
    for i, it in enumerate(items, start=1):
        seq = it.get("seq", str(i))
        part_no = str(it.get("part_no", "TBD")).upper()
        part_name = it.get("part_name", "")
        qty = int(it.get("qty", 1))
        category = it.get("category", "")
        material = it.get("material", "")
        mspec = it.get("material_spec", "")
        texture = it.get("texture", "")
        second_process = it.get("second_process", "")
        vol = float(it.get("volume_cm3", 0.0))
        mass = float(it.get("mass_g", 0.0))
        lmm = float(it.get("l_mm", 0.0) or 0.0)
        wmm = float(it.get("w_mm", 0.0) or 0.0)
        hmm = float(it.get("h_mm", 0.0) or 0.0)
        if lmm > 0 and wmm > 0 and hmm > 0:
            vol_display = f"{lmm:.0f} x {wmm:.0f} x {hmm:.0f}"
        else:
            vol_display = f"{vol:.2f}"

        img_b64 = it.get("preview_png_b64", "")
        img_html = (
            f"""<div class="imgbox"><img src="data:image/png;base64,{img_b64}" onload="_setHeight();" onerror="_setHeight();"/></div>"""
            if img_b64
            else "No Image"
        )

        rows.append([
            esc(seq),
            esc(part_no),
            esc(part_name),
            img_html,
            f"""<span class="mono">{qty}</span>""",
            esc(category),
            esc(material),
            esc(mspec),
            f"""<div class="ml">{esc_multiline(texture)}</div>""",
            f"""<div class="ml">{esc_multiline(second_process)}</div>""",
            f"""<span class="mono">{esc(vol_display)}</span>""",
            f"""<span class="mono">{mass:.2f}</span>""",
        ])

    _render_html_table("🧾 BOM（HLR Preview / 只讀顯示）", columns, rows, row_h, HEADER_HEIGHT_PX)


def render_tooling_readonly(items: list[dict]) -> None:
    filtered = [it for it in items if str(it.get('category','')).lower() != 'sub-assy']
    items = filtered
    row_h = max(140, int(IMAGE_WIDTH_PX * 1.55))

    columns = [
        ("#", "4%"),
        ("Part No.", "7%"),
        ("Tooling No.", "8%"),
        ("Part Name", "13%"),
        ("Preview", "10%"),
        ("Category", "10%"),
        ("Runner", "7%"),
        ("Tooling Life", "7%"),
        ("Cavities", "7%"),
        ("Texture", "10%"),
        ("Cavity Mat.", "7%"),
        ("Core Mat.", "7%"),
        ("Lifter", "6%"),
        ("Slider", "6%"),
    ]

    rows = []
    for i, it in enumerate(items, start=1):
        seq = it.get("seq", str(i))
        part_no = str(it.get("part_no", "TBD")).upper()
        tooling_no = it.get("tooling_no", "TBD") or "TBD"
        part_name = it.get("part_name", "")
        category = it.get("category", "")
        runner = it.get("runner", "N/A")
        tooling_life_k = it.get("tooling_life_k", "")
        cavities = it.get("cavities", "1")
        texture = it.get("texture", "")
        cav_mat = it.get("cavity_material", "")
        core_mat = it.get("core_material", "")
        lifter = it.get("lifter_qty", 0)
        slider = it.get("slider_qty", 0)

        img_b64 = it.get("preview_png_b64", "")
        img_html = (
            f"""<div class="imgbox"><img src="data:image/png;base64,{img_b64}" onload="_setHeight();" onerror="_setHeight();"/></div>"""
            if img_b64
            else "No Image"
        )

        rows.append([
            esc(seq),
            esc(part_no),
            esc(tooling_no),
            esc(part_name),
            img_html,
            esc(category),
            esc(runner),
            esc(str(tooling_life_k)),
            esc(str(cavities)),
            f"""<div class="ml">{esc_multiline(texture)}</div>""",
            esc(cav_mat),
            esc(core_mat),
            f"""<span class="mono">{int(lifter)}</span>""",
            f"""<span class="mono">{int(slider)}</span>""",
        ])

    _render_html_table("🧰 Tooling List（只讀顯示）", columns, rows, row_h, HEADER_HEIGHT_PX)


# ================= APP =================
st.set_page_config(page_title="STEP BOM + Tooling List", page_icon="📐", layout="wide")
ensure_state()

st.markdown(
    """
<style>
[data-testid="stDataFrame"] * { font-size: 12px !important; }
[data-testid="stDataFrame"] [role="columnheader"] * { font-size: 12px !important; }
</style>
""",
    unsafe_allow_html=True,
)

st.title("📐 STEP → BOM（HLR Preview + Inline Edit） + 🧰 Tooling List")

# meta
meta_cols = st.columns([1.2, 1.2, 2.6])
with meta_cols[0]:
    st.session_state["project"] = st.text_input("Project", st.session_state["project"])
with meta_cols[1]:
    st.session_state["engineer"] = st.text_input("Engineer", st.session_state["engineer"])
with meta_cols[2]:
    st.caption("建議：先填 Project / Engineer，再開始 New BOM 或 Load。")

st.markdown("---")

mode = st.radio("開始方式", ["New BOM", "繼續編輯（載入 bom.json）"], horizontal=True)

# --- Top bar layout (依照你貼圖的需求)：
# 左：New BOM 或 Load
# 中：Save BOM
# 右：Excel 輸出
bom_file = None
topbar = st.columns([2.2, 2.2, 2.2])

with topbar[0]:
    if mode == "New BOM":
        if st.button("🆕 建立新 BOM", use_container_width=True, disabled=st.session_state["new_bom_used"]):
            reset_bom()
            st.session_state["new_bom_used"] = True
            st.session_state["bom_context_ready"] = True
            st.success("已建立新 BOM")
            st.rerun()
    else:
        st.session_state["new_bom_used"] = False
        bom_file = st.file_uploader("📂 Load BOM（bom.json）", type=["json"], key="load_bom_top")

with topbar[1]:
    save_payload = {
        "meta": {
            "project": st.session_state["project"],
            "engineer": st.session_state["engineer"],
            "saved_at": datetime.now().isoformat(timespec="seconds"),
            "version": 9,
        },
        "items": st.session_state["bom"],
    }
    default_filename = (st.session_state["project"].strip() or "bom") + ".json"
    st.download_button(
        "💾 Save BOM (bom.json)",
        data=json.dumps(save_payload, ensure_ascii=False, indent=2),
        file_name=default_filename,
        mime="application/json",
        use_container_width=True,
        disabled=False,
    )

with topbar[2]:
    proj_name = _safe_filename(st.session_state.get("project", ""))
    if not _has_openpyxl():
        st.warning("Excel 匯出需要 openpyxl，請在目前的環境安裝：pip install openpyxl")
    elif st.session_state.get("bom"):
        bom_df_export = build_bom_export_df(st.session_state["bom"])
        tooling_df_export = build_tooling_export_df(st.session_state["bom"])

        bom_xlsx = df_to_excel_bytes(bom_df_export, sheet_name="ME_BOM")
        tooling_xlsx = df_to_excel_bytes(tooling_df_export, sheet_name="Tooling_List")

        st.download_button(
            "📄 Generate Project_ME_BOM.xlsx",
            data=bom_xlsx,
            file_name=f"{proj_name}_ME_BOM.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.download_button(
            "🧰 Generate Project_Tooling_List.xlsx",
            data=tooling_xlsx,
            file_name=f"{proj_name}_Tooling_List.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        st.info("先建立/載入 BOM 後才可輸出 Excel")

# Load
if mode == "繼續編輯（載入 bom.json）":
    if bom_file is not None:
        try:
            data = json.load(bom_file)
            meta = data.get("meta", {})
            items = data.get("items", [])
            if not isinstance(items, list):
                raise ValueError("bom.json 格式錯誤：items 不是 list")

            st.session_state["project"] = meta.get("project", st.session_state["project"])
            st.session_state["engineer"] = meta.get("engineer", st.session_state["engineer"])

            fixed = []
            for idx, it in enumerate(items, start=1):
                it.setdefault("id", str(uuid4()))
                it.setdefault("seq", str(idx))
                it.setdefault("part_no", "TBD")
                it["part_no"] = str(it.get("part_no", "TBD")).upper()
                it.setdefault("part_name", it.get("step_name", "") or "Unnamed")
                it.setdefault("qty", 1)

                it.setdefault("category", "plastic molding (single shot)")
                it.setdefault("material", "PC/ABS")
                it.setdefault("material_spec", "")
                it.setdefault("texture", "")

                # Sub-Assy 主階不需要 Resin Code / Texture：固定為 "-" 且不可修改
                if str(it.get("category","")).strip().lower() in {"sub-assy","sub assy","subassy"}:
                    it["material_spec"] = "-"
                    it["texture"] = "-"
                it.setdefault("second_process", "")

                if "density" not in it:
                    it["density"] = float(MATERIAL_DENSITY.get(it["material"], 1.0))
                it.setdefault("volume_cm3", 0.0)
                it.setdefault("mass_g", float(it["volume_cm3"]) * float(it["density"]))

                it.setdefault("tooling_no", "TBD")
                if not it.get("tooling_no"):
                    it["tooling_no"] = "TBD"

                it.setdefault("runner", "N/A")
                it.setdefault("tooling_life_k", "500K")
                if not str(it.get("tooling_life_k","")).strip():
                    it["tooling_life_k"] = "500K"
                it.setdefault("cavities", "1")

                # Auto-suggest cavities when missing/blank/default
                cur_cav = str(it.get("cavities", "")).strip()
                if cur_cav in ("", "-", "1"):
                    it["cavities"] = suggest_cavities_str(
                        it.get("category", ""),
                        float(it.get("l_mm", 0.0) or 0.0),
                        float(it.get("w_mm", 0.0) or 0.0),
                        float(it.get("h_mm", 0.0) or 0.0),
                        float(it.get("volume_cm3", it.get("step_volume_cm3", 0.0)) or 0.0),
                    )
                it.setdefault("cavity_material", "")
                it.setdefault("core_material", "")
                it.setdefault("lifter_qty", 0)
                it.setdefault("slider_qty", 0)

                fixed.append(recalc_item(it))

            st.session_state["bom"] = fixed
            st.session_state["bom_context_ready"] = True
            st.success(f"已載入 {len(st.session_state['bom'])} 筆料件，可續編。")
        except Exception as e:
            st.error(f"載入失敗：{e}")

st.markdown("---")


if not st.session_state["bom_context_ready"]:
    st.info("請先選擇 New BOM 或載入 bom.json 後，再新增料件（上傳 STEP）。")
else:
    st.subheader("➕ 新增料件")
    mode_cols = st.columns([1.2, 1.2, 3.6])
    with mode_cols[0]:
        if st.button("📦 Part Upload", use_container_width=True):
            st.session_state["add_mode"] = "part"
    with mode_cols[1]:
        if st.button("🧩 Sub-Assy Upload", use_container_width=True):
            st.session_state["add_mode"] = "subassy"
    with mode_cols[2]:
        if st.session_state["add_mode"] == "":
            st.caption("先選擇 Part Upload 或 Sub-Assy Upload，才會出現唯一的 STEP 上傳入口。")
        elif st.session_state["add_mode"] == "part":
            st.caption("模式：Part Upload（可一次多檔加入 BOM）")
        else:
            st.caption("模式：Sub-Assy Upload（先上傳主階，再上傳子階並回填主階 2nd Process / Assembly / Assembly）")

    st.markdown(
        """
<style>
div[data-testid="stFileUploaderFileList"] { display: none !important; }
div[data-testid="stFileUploader"] ul { display: none !important; }
div[data-testid="stFileUploader"] div[role="list"] { display: none !important; }
div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] { display: none !important; }
div[data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"] ~ * { display: none !important; }
</style>
""",
        unsafe_allow_html=True,
    )

    # ---------- Part Upload ----------
    if st.session_state["add_mode"] == "part":
        uploaded_steps = st.file_uploader(
            "Upload STEP (.step/.stp) — Part Upload",
            type=["step", "stp"],
            accept_multiple_files=True,
            key=st.session_state["uploader_key"],
        )

        # 收集上傳檔案到暫存清單（可刪除），避免 file_uploader 只能看不能刪
        if uploaded_steps:
            for uf in uploaded_steps:
                buf = uf.getbuffer()
                size_bytes = getattr(uf, "size", None) or len(buf)
                st.session_state["part_pending_files"].append({
                    "id": str(uuid4()),
                    "name": uf.name,
                    "size": int(size_bytes),
                    "bytes": bytes(buf),
                })
            # 清空 uploader 選擇，避免 rerun 重複加入同一批檔案
            st.session_state["uploader_key"] = str(uuid4())
            st.rerun()

        pending = st.session_state.get("part_pending_files", [])
        if pending:
            st.caption(f"已選擇 {len(pending)} 個檔案")
            with st.expander("查看已選檔名", expanded=False):
                for i, pf in enumerate(pending, start=1):
                    c1, c2, c3 = st.columns([0.08, 0.74, 0.18])
                    with c1:
                        st.write(f"{i}.")
                    with c2:
                        st.write(f"{pf['name']} ({pf['size']/1024:.1f} KB)")
                    with c3:
                        if st.button("✖ 刪除", key=f"rm_part_{pf['id']}"):
                            st.session_state["part_pending_files"] = [x for x in pending if x.get("id") != pf.get("id")]
                            st.rerun()


        add_col = st.columns([1.3, 2.7, 1.0])
        with add_col[0]:
            if st.button("✅ 解析並加入 BOM", use_container_width=True):
                if not pending:
                    st.warning("請先選擇 STEP 檔案")
                else:
                    progress = st.progress(0)
                    status = st.empty()
                    total = len(pending)
                    added = 0
                    failed = []
                    fail_msgs = []

                    with st.spinner("解析中…（STEP → Volume + HLR Preview）"):
                        for k, pf in enumerate(pending, start=1):
                            status.info(f"處理中：{k}/{total}  -  {pf['name']}")
                            progress.progress(int((k - 1) / total * 100))

                            with tempfile.NamedTemporaryFile(delete=False, suffix=".step") as f:
                                f.write(pf['bytes'])
                                step_path = f.name

                            try:
                                shape = load_shape_from_step(step_path)
                                vol_mm3 = compute_volume_mm3(shape)
                                vol_cm3 = vol_mm3 / 1000.0
                                dx, dy, dz = compute_bbox_dims_mm(shape)
                                bbox_vol_cm3 = (dx * dy * dz) / 1000.0

                                png_bytes = make_hlr_png_bytes(shape, dpi=200, fig_size=4, sample_n=60)
                                png_b64 = png_bytes_to_b64(png_bytes)

                                default_name = str(pf['name']).rsplit(".", 1)[0]
                                next_seq = get_next_main_seq(st.session_state["bom"])

                                default_mat = "PC/ABS"
                                default_den = float(MATERIAL_DENSITY.get(default_mat, 1.0))

                                item = {
                                    "id": str(uuid4()),
                                    "seq": str(next_seq),
                                    "part_no": "TBD",
                                    "part_name": default_name,
                                    "step_name": pf['name'],
                                    "qty": 1,
                                    "category": "plastic molding (single shot)",
                                    "material": default_mat,
                                    "material_spec": "",
                                    "texture": "",
                                    "second_process": "",
                                    "density": default_den,
                                    "step_volume_cm3": float(vol_cm3),
                                    "l_mm": float(dx),
                                    "w_mm": float(dy),
                                    "h_mm": float(dz),
                                    "bbox_volume_cm3": float(bbox_vol_cm3),
                                    "volume_cm3": float(vol_cm3),
                                    "mass_g": float(vol_cm3) * default_den,
                                    "preview_png_b64": png_b64,
                                    "tooling_no": "TBD",
                                    "runner": "N/A",
                                    "tooling_life_k": "500K",
                                                                        "cavities": suggest_cavities_str("plastic molding (single shot)", float(dx), float(dy), float(dz), float(vol_cm3)),
                                    "cavity_material": "",
                                    "core_material": "",
                                    "lifter_qty": 0,
                                    "slider_qty": 0,
                                }
                                st.session_state["bom"].append(recalc_item(item))
                                added += 1
                            except Exception as e:
                                failed.append((pf['name'], str(e)))
                                fail_msgs.append(f"{pf['name']}: {e}")
                                st.error(f"新增失敗：{pf['name']} / {e}")

                        progress.progress(100)
                        status.success(f"完成：已新增 {added}/{total} 筆")
                        st.session_state["last_add_report"] = {"added": added, "total": total, "failed": failed}
                        st.session_state["last_add_errors"] = fail_msgs

                    if added:
                        st.session_state["bom_version"] += 1
                        st.session_state["tooling_version"] += 1
                        st.session_state["part_pending_files"] = []
                        st.session_state["uploader_key"] = str(uuid4())
                        st.rerun()

        with add_col[1]:
            st.caption("解析加入後，上傳區會自動清空，避免重複加入。")
        with add_col[2]:
            st.caption(f"Total: {len(st.session_state['bom'])}")

    # ---------- Sub-Assy Upload ----------
    if st.session_state["add_mode"] == "subassy":
        st.markdown("#### 1) 上傳 Sub-Assy 主階（單檔）")
        parent_up = st.file_uploader(
            "Upload STEP (.step/.stp) — Sub-Assy 主階",
            type=["step", "stp"],
            accept_multiple_files=False,
            key="subassy_parent_uploader_" + st.session_state["subassy_uploader_key"],
        )

        # 顯示已選檔名（避免全域 CSS 隱藏 uploader 的檔名清單）
        if parent_up is not None:
            st.session_state["subassy_parent_file"] = parent_up.name
        picked_parent = st.session_state.get("subassy_parent_file", "")
        if picked_parent:
            st.caption(f"已選檔案：{picked_parent}")


        c_parent = st.columns([1.2, 3.8])
        with c_parent[0]:
            if st.button("✅ 建立 Sub-Assy 主階", use_container_width=True):
                if parent_up is None:
                    st.warning("請先選擇主階 STEP 檔案")
                else:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".step") as f:
                        f.write(parent_up.getbuffer())
                        step_path = f.name
                    try:
                        shape = load_shape_from_step(step_path)
                        vol_mm3 = compute_volume_mm3(shape)
                        vol_cm3 = vol_mm3 / 1000.0
                        dx, dy, dz = compute_bbox_dims_mm(shape)
                        bbox_vol_cm3 = (dx * dy * dz) / 1000.0
                        png_bytes = make_hlr_png_bytes(shape, dpi=200, fig_size=4, sample_n=60)
                        png_b64 = png_bytes_to_b64(png_bytes)

                        default_name = parent_up.name.rsplit(".", 1)[0]
                        parent_seq = get_next_main_seq(st.session_state["bom"])
                        item = {
                            "id": str(uuid4()),
                            "seq": parent_seq,
                            "part_no": "TBD",
                            "part_name": default_name,
                            "step_name": parent_up.name,
                            "qty": 1,
                            "category": "sub-assy",
                            "material": "PC/ABS",
                            "material_spec": "-",
                            "texture": "-",
                            "second_process": "",
                            "density": float(MATERIAL_DENSITY.get("PC/ABS", 1.0)),
                            "step_volume_cm3": float(vol_cm3),
                            "l_mm": float(dx),
                            "w_mm": float(dy),
                            "h_mm": float(dz),
                            "bbox_volume_cm3": float(bbox_vol_cm3),
                            "volume_cm3": float(vol_cm3),
                            "mass_g": float(vol_cm3) * float(MATERIAL_DENSITY.get("PC/ABS", 1.0)),
                            "preview_png_b64": png_b64,
                            "tooling_no": "TBD",
                            "runner": "N/A",
                            "tooling_life_k": "500K",
                                                                "cavities": suggest_cavities_str("sub-assy", float(dx), float(dy), float(dz), float(vol_cm3)),
                            "cavity_material": "",
                            "core_material": "",
                            "lifter_qty": 0,
                            "slider_qty": 0,
                        }
                        st.session_state["bom"].append(recalc_item(item))
                        st.session_state["subassy_parent_id"] = item["id"]
                        st.session_state["subassy_parent_seq"] = parent_seq
                        st.session_state["subassy_child_n"] = 0
                        st.session_state["subassy_child_meta"] = {}
                        st.session_state["subassy_child_uploaded"] = {}
                        st.session_state["bom_version"] += 1
                        st.session_state["uploader_key"] = str(uuid4())
                        st.session_state["subassy_uploader_key"] = str(uuid4())
                        st.success(f"已建立 Sub-Assy 主階：{parent_seq}")
                        st.rerun()
                    except Exception as e:
                        st.error(f"建立主階失敗：{e}")

        with c_parent[1]:
            if st.session_state.get("subassy_parent_id"):
                st.caption(f"目前主階：Seq={st.session_state.get('subassy_parent_seq')}")

        if st.session_state.get("subassy_parent_id"):
            st.markdown("#### 2) 設定子階數量並上傳（每個子階各自一個 STEP）")
            st.session_state["subassy_child_n"] = st.number_input(
                "子階數量",
                min_value=1,
                max_value=20,
                step=1,
                value=int(st.session_state.get("subassy_child_n") or 1),
            )

            parent_seq = st.session_state["subassy_parent_seq"]
            n = int(st.session_state["subassy_child_n"])
            for i in range(1, n + 1):
                child_seq = f"{parent_seq}-{i}"

                # init meta
                meta = st.session_state["subassy_child_meta"].get(child_seq, {"method": ASSEMBLY_OPTIONS[0], "qty": 1, "other": ""})
                up_ok = bool(st.session_state["subassy_child_uploaded"].get(child_seq, False))

                st.markdown("---")
                cols = st.columns([1.55, 0.9, 1.35, 1.2, 1.8])
                with cols[0]:
                    up = st.file_uploader(
                        f"{child_seq} — STEP",
                        type=["step", "stp"],
                        accept_multiple_files=False,
                        key=f"child_uploader_{child_seq}_{st.session_state['subassy_uploader_key']}",
                    )
                    if up is not None:
                        st.session_state["subassy_child_uploaded"][child_seq] = True
                        st.session_state["subassy_child_meta"][child_seq] = meta
                        st.session_state["subassy_child_files"][child_seq] = up.name

                    picked_name = st.session_state.get("subassy_child_files", {}).get(child_seq, "")
                    if picked_name:
                        st.caption(f"已選檔案：{picked_name}")

                with cols[1]:
                    st.text_input("子階代號", value=child_seq, disabled=True, key=f"child_code_{child_seq}")
                with cols[2]:
                    meta["method"] = st.selectbox(
                        "組裝方式",
                        ASSEMBLY_OPTIONS,
                        index=ASSEMBLY_OPTIONS.index(meta.get("method", ASSEMBLY_OPTIONS[0])) if meta.get("method") in ASSEMBLY_OPTIONS else 0,
                        key=f"child_method_{child_seq}",
                    )
                with cols[3]:
                    meta["qty"] = st.number_input("數量(處)", min_value=1, step=1, value=int(meta.get("qty", 1)), key=f"child_qty_{child_seq}")
                with cols[4]:
                    meta["other"] = st.text_input("其他（若選其他）", value=meta.get("other",""), key=f"child_other_{child_seq}")

                st.session_state["subassy_child_meta"][child_seq] = meta

                if up_ok:
                    st.caption("✅ 已選擇 STEP")
                else:
                    st.caption("⬜ 尚未選擇 STEP")

                # store uploaded file bytes temporarily in session? (streamlit keeps by key; we will re-read via st.session_state)
                # We'll parse on Apply to avoid heavy ops per change.

            st.markdown("---")
            a1, a2, a3 = st.columns([1.6, 1.6, 2.8])
            with a1:
                if st.button("✅ 解析並加入 BOM", use_container_width=True):
                    # gather uploaded child files from session_state by uploader keys
                    missing = []
                    child_files = []
                    for i in range(1, n + 1):
                        child_seq = f"{parent_seq}-{i}"
                        key = f"child_uploader_{child_seq}_{st.session_state['subassy_uploader_key']}"
                        uf = st.session_state.get(key, None)
                        if uf is None:
                            missing.append(child_seq)
                        else:
                            child_files.append((child_seq, uf))

                    if missing:
                        st.warning("以下子階尚未上傳 STEP：" + ", ".join(missing))
                    else:
                        progress = st.progress(0)
                        status = st.empty()
                        total = len(child_files)
                        added = 0
                        fail = False

                        with st.spinner("處理子階 STEP…"):
                            for k, (child_seq, uf) in enumerate(child_files, start=1):
                                status.info(f"處理中：{k}/{total}  -  {child_seq}  -  {uf.name}")
                                progress.progress(int((k - 1) / total * 100))

                                with tempfile.NamedTemporaryFile(delete=False, suffix=".step") as f:
                                    f.write(uf.getbuffer())
                                    step_path = f.name

                                try:
                                    shape = load_shape_from_step(step_path)
                                    vol_mm3 = compute_volume_mm3(shape)
                                    vol_cm3 = vol_mm3 / 1000.0
                                    dx, dy, dz = compute_bbox_dims_mm(shape)
                                    bbox_vol_cm3 = (dx * dy * dz) / 1000.0
                                    png_bytes = make_hlr_png_bytes(shape, dpi=200, fig_size=4, sample_n=60)
                                    png_b64 = png_bytes_to_b64(png_bytes)

                                    default_name = uf.name.rsplit(".", 1)[0]

                                    default_mat = "PC/ABS"
                                    default_den = float(MATERIAL_DENSITY.get(default_mat, 1.0))

                                    item = {
                                        "id": str(uuid4()),
                                        "seq": child_seq,
                                        "part_no": "TBD",
                                        "part_name": default_name,
                                        "step_name": uf.name,
                                        "qty": 1,
                                        "category": "sub-assy child",
                                        "material": default_mat,
                                        "material_spec": "",
                                        "texture": "",
                                        "second_process": "",
                                        "density": default_den,
                                        "step_volume_cm3": float(vol_cm3),
                                        "l_mm": float(dx),
                                        "w_mm": float(dy),
                                        "h_mm": float(dz),
                                        "bbox_volume_cm3": float(bbox_vol_cm3),
                                    "volume_cm3": float(vol_cm3),
                                        "mass_g": float(vol_cm3) * default_den,
                                        "preview_png_b64": png_b64,
                                        "tooling_no": "TBD",
                                        "runner": "N/A",
                                        "tooling_life_k": "500K",
                                                                            "cavities": suggest_cavities_str("sub-assy child", float(dx), float(dy), float(dz), float(vol_cm3)),
                                        "cavity_material": "",
                                        "core_material": "",
                                        "lifter_qty": 0,
                                        "slider_qty": 0,
                                    }
                                    st.session_state["bom"].append(recalc_item(item))
                                    added += 1
                                except Exception as e:
                                    fail = True
                                    st.error(f"子階新增失敗：{child_seq} / {e}")

                        progress.progress(100)
                        if fail:
                            st.warning("子階處理完成，但有失敗項目（請看上方錯誤）。")
                        else:
                            status.success(f"子階新增完成：{added}/{total}")

                        # write back parent second_process
                        pid = st.session_state["subassy_parent_id"]
                        parent = None
                        for it in st.session_state["bom"]:
                            if it.get("id") == pid:
                                parent = it
                                break
                        if parent is not None:
                            lines = []
                            for i in range(1, n + 1):
                                child_seq = f"{parent_seq}-{i}"
                                meta = st.session_state["subassy_child_meta"].get(child_seq, {})
                                method = str(meta.get("method", "")).strip()
                                if method == "其他":
                                    method = str(meta.get("other","")).strip() or "其他"
                                qty = int(meta.get("qty", 1))
                                lines.append(f"{method}{child_seq} {qty}處")
                            new_text = "\n".join(lines).strip()
                            if parent.get("second_process","").strip():
                                parent["second_process"] = (parent["second_process"].rstrip() + "\n" + new_text).strip()
                            else:
                                parent["second_process"] = new_text
                            st.success("已回填主階 2nd Process / Assembly / Assembly")

                        st.session_state["bom_version"] += 1
                        st.session_state["tooling_version"] += 1
                        # reset subassy state for next use
                        st.session_state["subassy_parent_id"] = ""
                        st.session_state["subassy_parent_seq"] = ""
                        st.session_state["subassy_child_n"] = 0
                        st.session_state["subassy_child_meta"] = {}
                        st.session_state["subassy_child_uploaded"] = {}
                        st.session_state["subassy_uploader_key"] = str(uuid4())
                        st.session_state["uploader_key"] = str(uuid4())
                        st.rerun()
            with a2:
                if st.button("↩️ 取消 Sub-Assy", use_container_width=True):
                    st.session_state["subassy_parent_id"] = ""
                    st.session_state["subassy_parent_seq"] = ""
                    st.session_state["subassy_child_n"] = 0
                    st.session_state["subassy_child_meta"] = {}
                    st.session_state["subassy_child_uploaded"] = {}
                    st.session_state["subassy_uploader_key"] = str(uuid4())
                    st.session_state["add_mode"] = ""
                    st.rerun()
            with a3:
                st.caption("解析加入後會新增子階料件，並把主階 2nd Process / Assembly 自動填入「組裝方式 + 子階代號 + 數量」。")

st.markdown("---")

if not st.session_state["bom"]:
    st.info("目前沒有料件。")
    st.stop()

for it in st.session_state["bom"]:
    it.setdefault("density", float(MATERIAL_DENSITY.get(it.get("material", "PC/ABS"), 1.0)))
    it.setdefault("texture", "")
    it.setdefault("second_process", "")
    # Sub-Assy 主階不需要 Resin Code / Texture：固定為 "-" 且不可修改
    if str(it.get("category","")).strip().lower() in {"sub-assy","sub assy","subassy"}:
        it["material_spec"] = "-"
        it["texture"] = "-"
    it.setdefault("tooling_no", "TBD")
    if not it.get("tooling_no"):
        it["tooling_no"] = "TBD"
    if not str(it.get("tooling_life_k","")).strip():
        it["tooling_life_k"] = "500K"
    it.setdefault("runner", "N/A")
    it.setdefault("cavities", "1")
    recalc_item(it)

id2item = {it["id"]: it for it in st.session_state["bom"]}

render_bom_readonly(st.session_state["bom"])

st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)

st.markdown('<div style="height:22px"></div>', unsafe_allow_html=True)
st.subheader("✏️ BOM Inline Edit（BOM 欄位）")

rows = []
for it in st.session_state["bom"]:
    rows.append({
        "id": it["id"],
        "Seq": it.get("seq", ""),
        "Part No.": str(it.get("part_no", "TBD")).upper(),
        "Part Name": it.get("part_name", ""),
        "QTY": int(it.get("qty", 1)),
        "Category": it.get("category", "plastic molding (single shot)"),
        "Material": it.get("material", "PC/ABS"),
        "Resin Code": it.get("material_spec", ""),
        "Texture": it.get("texture", ""),
        "Density": float(it.get("density", 0.0)),
        "Vol (cm³)": float(it.get("volume_cm3", it.get("step_volume_cm3", 0.0)) or 0.0),
        "Del": False,
    })

df_all = pd.DataFrame(rows)
id_series = df_all["id"].copy()
df_show = df_all.drop(columns=["id"])

edited = st.data_editor(
    df_show,
    use_container_width=True,
    hide_index=True,
    key=f"bom_editor_{st.session_state['bom_version']}",
    disabled=["Vol (cm³)"],
    column_config={
        "Seq": st.column_config.TextColumn("Seq", width="small"),
        "Part No.": st.column_config.TextColumn("Part No.", width="small"),
        "Part Name": st.column_config.TextColumn("Part Name", width="medium"),
        "QTY": st.column_config.NumberColumn("QTY", min_value=1, step=1, width="small"),
        "Category": st.column_config.SelectboxColumn("Category", options=CATEGORY_OPTIONS, width="medium"),
        "Material": st.column_config.SelectboxColumn("Material", options=list(MATERIAL_DENSITY.keys()), width="small"),
        "Resin Code": st.column_config.TextColumn("Resin Code", width="medium"),
        "Texture": st.column_config.TextColumn("Texture", width="medium"),
        "Density": st.column_config.NumberColumn("Density", format="%.3f", width="small"),
        "Vol (cm³)": st.column_config.NumberColumn("Vol (cm³)", format="%.3f", width="small"),
        "Del": st.column_config.CheckboxColumn("Del", width="small"),
    },
)

bbtn = st.columns([1.2, 1.6, 1.6, 3.6])
with bbtn[0]:
    apply_bom = st.button("✅ Apply BOM", use_container_width=True)
with bbtn[1]:
    del_bom = st.button("🗑️ Delete Checked", use_container_width=True)
with bbtn[2]:
    auto_den = st.button("↺ Auto Density from Material", use_container_width=True)
with bbtn[3]:
    st.caption("2nd Process / Assembly 多行請用下方編輯器（Enter 可換行）。")

st.markdown("### 🧴 2nd Process / Assembly（多行編輯）")
labels = []
label2id = {}
for it in st.session_state["bom"]:
    label = f"{it.get('seq','')} | {it.get('part_no','TBD')} | {it.get('part_name','')}"
    labels.append(label)
    label2id[label] = it["id"]

if labels:
    chosen_label = st.selectbox("選擇要編輯的料件", labels, index=0)
    chosen_id = label2id[chosen_label]

    cur = id2item[chosen_id].get("second_process", "")
    if st.session_state["process_text"] != cur:
        st.session_state["process_text"] = cur

    st.session_state["process_text"] = st.text_area(
        "2nd Process / Assembly（可多行，按 Enter 換行）",
        st.session_state["process_text"],
        height=170,
    )

    c1, c2 = st.columns([1.2, 3.8])
    with c1:
        if st.button("💾 Save 2nd Process / Assembly", use_container_width=True):
            id2item[chosen_id]["second_process"] = st.session_state["process_text"]
            st.success("已更新 2nd Process / Assembly")
            st.rerun()
    with c2:
        st.caption("（註）st.data_editor 不支援格子內多行，所以用 TextArea。")

if del_bom:
    del_rows = edited.loc[edited["Del"] == True].index.tolist()
    del_ids = set(id_series.iloc[i] for i in del_rows)
    st.session_state["bom"] = [it for it in st.session_state["bom"] if it["id"] not in del_ids]
    st.rerun()

if auto_den:
    for idx, r in edited.iterrows():
        pid = id_series.iloc[idx]
        it = id2item.get(pid)
        if not it:
            continue
        mat = str(r["Material"])
        it["density"] = float(MATERIAL_DENSITY.get(mat, it.get("density", 1.0)))
        recalc_item(it)
    st.success("Density 已依 Material 回復為預設值")
    st.rerun()

if apply_bom:
    for idx, r in edited.iterrows():
        pid = id_series.iloc[idx]
        it = id2item.get(pid)
        if not it:
            continue

        old_mat = it.get("material", "PC/ABS")
        new_mat = str(r["Material"])
        old_den = float(it.get("density", float(MATERIAL_DENSITY.get(old_mat, 1.0))))
        new_den_user = float(r["Density"])

        it["seq"] = str(r["Seq"]).strip()
        it["part_no"] = str(r["Part No."]).strip().upper() or "TBD"
        it["part_name"] = str(r["Part Name"])
        it["qty"] = int(r["QTY"]) if str(r["QTY"]).strip() != "" else 1
        it["category"] = str(r["Category"])
        it["material"] = new_mat
        # Sub-Assy 主階 Resin Code / Texture 固定為 "-"（不可修改）
        if str(it.get("category","")).strip().lower() in {"sub-assy","sub assy","subassy"}:
            it["material_spec"] = "-"
            it["texture"] = "-"
        else:
            it["material_spec"] = str(r["Resin Code"])
            it["texture"] = str(r["Texture"])

        # --- Cavities suggestion (trigger after Category is chosen) ---
        # Only auto-suggest when cavities is still blank/default ("", "-", "1").
        # Users can always override later in Tooling editor.
        cur_cav = str(it.get("cavities", "")).strip()
        if cur_cav in ("", "-", "1"):
            it["cavities"] = suggest_cavities_str(
                it.get("category", ""),
                float(it.get("l_mm", 0.0) or 0.0),
                float(it.get("w_mm", 0.0) or 0.0),
                float(it.get("h_mm", 0.0) or 0.0),
                float(it.get("volume_cm3", it.get("step_volume_cm3", 0.0)) or 0.0),
            )

        # Volume shown in editor should keep original STEP volume formula (cm^3)
        it["volume_cm3"] = float(r.get("Vol (cm³)", it.get("volume_cm3", it.get("step_volume_cm3", 0.0))) or 0.0)

        auto_den_new = float(MATERIAL_DENSITY.get(new_mat, old_den))
        if (new_mat != old_mat) and abs(new_den_user - old_den) < 1e-9:
            it["density"] = auto_den_new
        else:
            it["density"] = new_den_user

        recalc_item(it)

    st.success("已套用 BOM 更新（Tooling List 會自動連動）")
    st.session_state["bom_version"] += 1
    st.session_state["tooling_version"] += 1
    st.rerun()


# ===== Tooling List =====
st.markdown("---")
render_tooling_readonly(st.session_state["bom"])

st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)

st.markdown('<div style="height:22px"></div>', unsafe_allow_html=True)
st.subheader("✏️ Tooling List Inline Edit（模具欄位）")

tool_rows = []
for it in st.session_state["bom"]:
    if str(it.get("category","")).lower() == "sub-assy":
        continue
    tool_rows.append({
        "id": it["id"],
        "Seq": it.get("seq", ""),
        "Part No.": str(it.get("part_no", "TBD")).upper(),
        "Tooling No.": it.get("tooling_no", "TBD") or "TBD",
        "Part Name": it.get("part_name", ""),
        "Runner": it.get("runner", "N/A"),
        "Tooling Life": it.get("tooling_life_k", ""),
        "Cavities": it.get("cavities", "1"),
        "Cavity Mat.": it.get("cavity_material", ""),
        "Core Mat.": it.get("core_material", ""),
        "Lifter": int(it.get("lifter_qty", 0)),
        "Slider": int(it.get("slider_qty", 0)),
    })

TOOL_EDITOR_COLS = [
    "id", "Seq", "Part No.", "Tooling No.", "Part Name",
    "Runner", "Tooling Life", "Cavities",
    "Cavity Mat.", "Core Mat.", "Lifter", "Slider"
]

tool_df_all = pd.DataFrame(tool_rows, columns=TOOL_EDITOR_COLS)

if tool_df_all.empty:
    st.info("目前沒有需要 Tooling 的料件（Sub-Assy 不列入 Tooling List）。")
else:
    tool_id_series = tool_df_all["id"].copy()
    tool_df_show = tool_df_all.drop(columns=["id"])

    tool_edited = st.data_editor(
        tool_df_show,
        use_container_width=True,
        hide_index=True,
        key=f"tool_editor_{st.session_state['tooling_version']}",
        disabled=["Part Name"],
        column_config={
            "Seq": st.column_config.TextColumn("Seq", width="small"),
            "Part No.": st.column_config.TextColumn("Part No.", width="small"),
            "Tooling No.": st.column_config.TextColumn("Tooling No.", width="small"),
            "Part Name": st.column_config.TextColumn("Part Name", width="medium"),
            "Runner": st.column_config.SelectboxColumn("Runner", options=RUNNER_OPTIONS, width="small"),
            "Tooling Life": st.column_config.TextColumn("Tooling Life", width="small"),
            "Cavities": st.column_config.TextColumn("Cavities", width="small"),
            "Cavity Mat.": st.column_config.SelectboxColumn("Cavity Mat.", options=TOOLING_MATERIAL_OPTIONS, width="small"),
            "Core Mat.": st.column_config.SelectboxColumn("Core Mat.", options=TOOLING_MATERIAL_OPTIONS, width="small"),
            "Lifter": st.column_config.NumberColumn("Lifter", min_value=0, step=1, width="small"),
            "Slider": st.column_config.NumberColumn("Slider", min_value=0, step=1, width="small"),
        },
    )

    tbtn = st.columns([1.4, 4.6])
    with tbtn[0]:
        apply_tool = st.button("✅ Apply Tooling", use_container_width=True)
    with tbtn[1]:
        st.caption("Tooling List 與 BOM 同一份資料；Category/Texture/Preview 會自動帶入並連動。")

    if apply_tool:
        for idx, r in tool_edited.iterrows():
            pid = tool_id_series.iloc[idx]
            it = id2item.get(pid)
            if not it:
                continue
            it["seq"] = str(r["Seq"]).strip()
            it["part_no"] = str(r["Part No."]).strip().upper() or "TBD"
            tn = str(r["Tooling No."]).strip()
            it["tooling_no"] = tn if tn else "TBD"
            it["runner"] = str(r["Runner"])
            it["tooling_life_k"] = str(r["Tooling Life"]).strip()
            it["cavities"] = (str(r["Cavities"]).strip() or "1")
            it["cavity_material"] = str(r["Cavity Mat."])
            it["core_material"] = str(r["Core Mat."])
            it["lifter_qty"] = int(r["Lifter"]) if str(r["Lifter"]).strip() != "" else 0
            it["slider_qty"] = int(r["Slider"]) if str(r["Slider"]).strip() != "" else 0

        st.success("已套用 Tooling 更新（BOM 會自動連動）")
        st.session_state["tooling_version"] += 1
        st.session_state["bom_version"] += 1
        st.rerun()

st.markdown("---")
total_weight = sum(float(x.get("mass_g", 0.0)) for x in st.session_state["bom"])
st.write(
    f"**Project:** {st.session_state['project'] or '-'}    |    "
    f"**Engineer:** {st.session_state['engineer'] or '-'}    |    "
    f"**Total parts:** {len(st.session_state['bom'])}    |    "
    f"**Total weight (g):** {total_weight:.2f}"
)
st.caption("bom.json 內含縮圖 base64 + 計算結果；載入可直接續編。新增料件才需要上傳新的 STEP。")
